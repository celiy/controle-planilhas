from tkinter import *
import xlsxwriter
from openpyxl import load_workbook
import openpyxl
import os

def arquivo():

    nome = insnome.get()
    nome = nome.upper()
    oqfeito = insoqfeito.get("1.0", END).strip()
    data = insdata.get("1.0", END).strip()

    file_path = 'planilha.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        worksheet = workbook.active
    else:
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Nome')
        worksheet.write('B1', 'Serviço Prestado')
        worksheet.write('C1', 'Data')
        workbook.close()
        workbook = load_workbook(file_path)
        worksheet = workbook.active

    existing_file = 'planilha.xlsx'

    #Adicionar as informações na planilha
    new_data = [[nome, oqfeito, data]]
    wb = load_workbook(existing_file)
    ws = wb.active
    for row in new_data:
        ws.append(row)
    wb.save(existing_file)

    workbook.close()

    insnome.delete(0, END)
    insoqfeito.delete("1.0", END)

def mostrarplanilha():

    def planilha():
        nome_procurado = nome_entry.get().strip()
        nome_procurado = nome_procurado.upper()
        dataframe = load_workbook("planilha.xlsx")
        dataframe1 = dataframe.active

        text_area.delete("1.0", END)

        for row in dataframe1.iter_rows(values_only=True):
            if len(nome_procurado) > 0:
                if nome_procurado in row:
                    row_text = "\t".join([str(cell) for cell in row])
                    text_area.insert(END, row_text + "\n")
            else:
                row_text = "\t".join([str(cell) for cell in row])
                text_area.insert(END, row_text + "\n")

    janela1 = Tk()
    janela1.title("Mostrar Planilha")

    nome_label = Label(janela1, text="Pesquisar nome:")
    nome_label.pack(pady=(12,0))

    nome_entry = Entry(janela1)
    nome_entry.pack()

    btn_mostrar = Button(janela1, text="Mostrar Planilha", command=planilha)
    btn_mostrar.pack(pady=(6,6))

    text_area = Text(janela1, wrap=NONE)
    text_area.pack(expand=True, fill=BOTH)

    scroll_x = Scrollbar(janela1, orient=HORIZONTAL, command=text_area.xview)
    scroll_x.pack(side=BOTTOM, fill=X)
    text_area.config(xscrollcommand=scroll_x.set)

    scroll_y = Scrollbar(janela1, orient=VERTICAL, command=text_area.yview)
    scroll_y.pack(side=RIGHT, fill=Y)
    text_area.config(yscrollcommand=scroll_y.set)

    janela1.mainloop()

janela = Tk()
janela.title("Controle de planilhas")

txtnome = Label(janela, text="Nome do cliente:").grid(row=0,column=0,pady=(12,0))
insnome = Entry(janela, width=25)
insnome.grid(row=1,column=0)

txtdataent = Label(janela, text="Data entrada:").grid(row=2,column=0,pady=(6,0))
txtinstr = Label(janela, text="hora-dia/mes/ano").grid(row=3,column=0)
insdata = Text(janela, width=15, height=1)
insdata.grid(row=4,column=0)
insdata.insert("1.0", "21-01/01/2024")

def clear_text(event):
    insdata.delete("1.0", END)

insdata.bind("<FocusIn>", clear_text)

txtoqfeito = Label(janela, text="Serviço prestado:").grid(row=6,column=0,pady=(6,0))
insoqfeito = Text(janela, width=30, height=10, pady=5)
insoqfeito.grid(row=7,column=0,padx=6)

submit = Button(janela, text="Salvar", command=arquivo, padx=5)
submit.grid(row=8,column=0,pady=(6,6))

mostrarplanilha()

janela.mainloop()